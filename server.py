from flask import Flask, request, send_file, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
import openpyxl.utils
from openpyxl.styles import numbers, Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.views import Pane
from copy import copy
import io
import base64
from datetime import datetime
import json
import traceback
import os
import logging
import gc
from logging.handlers import RotatingFileHandler
import urllib.parse
import db
import allocator
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# ============================================================
# НАСТРОЙКА ЛОГГИРОВАНИЯ В ФАЙЛЫ
# ============================================================
LOG_DIR = '/var/log/financial-analyzer'
os.makedirs(LOG_DIR, exist_ok=True)

# Корневой логгер
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)

# Формат логов
log_format = logging.Formatter(
    '[%(asctime)s] %(levelname)-8s %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Ротируемый файл (10 МБ, 5 файлов)
file_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, 'app.log'),
    maxBytes=10 * 1024 * 1024,
    backupCount=5,
    encoding='utf-8'
)
file_handler.setFormatter(log_format)
root_logger.addHandler(file_handler)

# Также пишем в stdout (для gunicorn/systemd)
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_format)
root_logger.addHandler(console_handler)

logger = logging.getLogger('financial_analyzer')

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB max file size

# ============================================================
# RATE LIMITING
# ============================================================
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["60 per minute"],
    storage_uri="memory://"
)

# ============================================================
# ИСПРАВЛЕНИЕ CORS - разрешаем запросы с любых источников
# ВАЖНО: expose_headers должен включать 'X-Filial-Data' для передачи данных филиалов
# ============================================================
CORS(app,
     expose_headers=['X-Filial-Data', 'X-Allocation-Summary'],
     resources={r'/*': {'origins': '*'}})

# ============================================================
# НАСТРОЙКИ КОДИРОВКИ ДЛЯ КИРИЛЛИЦЫ
# ============================================================
app.config['JSON_AS_ASCII'] = False
app.config['JSON_SORT_KEYS'] = False

# ============================================================
# МАРШРУТЫ ДЛЯ РАЗДАЧИ СТАТИКИ (фронтенд через HTTP)
# ============================================================
BASE_DIR = '/opt/financial-analyzer'

@app.route('/')
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/lib/<path:filename>')
def serve_lib(filename):
    return send_from_directory(os.path.join(BASE_DIR, 'lib'), filename)

@app.route('/<path:filename>')
def serve_static(filename):
    if '..' in filename or filename.startswith('/'):
        return jsonify({'error': 'Invalid path'}), 400
    return send_from_directory(BASE_DIR, filename)

# ============================================================
# ГЛОБАЛЬНЫЙ ОБРАБОТЧИК ДЛЯ УСТАНОВКИ КОДИРОВКИ
# ============================================================
@app.after_request
def add_charset_header(response):
    if response.content_type and response.content_type.startswith('application/json'):
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response

# ============================================================
# HEALTH CHECK ЭНДПОИНТ
# ============================================================
@app.route('/api/health', methods=['GET'])
def api_health():
    """Проверка работоспособности сервера и доступности БД."""
    db_ok = db.check_health()
    status_code = 200 if db_ok else 503
    return jsonify({
        'status': 'ok' if db_ok else 'degraded',
        'database': 'connected' if db_ok else 'unavailable',
        'timestamp': datetime.now().isoformat()
    }), status_code

# Список целевых контрагентов (используется только для фильтрации документов)
TARGET_CONTRAGENTS = ['ВАНКОРНЕФТЬ АО', 'РН-Ванкор ООО']

# Колонки (1‑индексация Excel)
COLUMNS = {
    'DOCUMENT_NAME': 1,      # A
    'DEBT_AMOUNT': 12,       # L
    'UNSIGNED_DEBT': 13,     # M
    'OVERDUE': 15,           # O
    'UNSIGNED_OVERDUE': 16,  # P
    'DAYS': 18,              # R
    'OUR_DEBT': 19,          # S
    'NOT_OVERDUE': 20,       # T
    'INTERVAL_1_15': 21,     # U
    'INTERVAL_16_29': 22,    # V
    'INTERVAL_30_89': 23,    # W
    'INTERVAL_90_179': 24,   # X
    'INTERVAL_180_PLUS': 25, # Y
}

SUM_COLUMNS = [
    COLUMNS['DEBT_AMOUNT'], COLUMNS['UNSIGNED_DEBT'],
    COLUMNS['OVERDUE'], COLUMNS['UNSIGNED_OVERDUE'],
    COLUMNS['NOT_OVERDUE'], COLUMNS['INTERVAL_1_15'],
    COLUMNS['INTERVAL_16_29'], COLUMNS['INTERVAL_30_89'],
    COLUMNS['INTERVAL_90_179'], COLUMNS['INTERVAL_180_PLUS'],
]

NUMERIC_COLUMNS = SUM_COLUMNS + [COLUMNS['DAYS']]

def is_cell_merged(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False

def get_cell_to_write(ws, row, col):
    if not is_cell_merged(ws, row, col):
        return ws.cell(row=row, column=col)
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return ws.cell(row=row, column=col)

def safe_set_number_format(ws, row, col, value):
    if row <= 13:
        return
    cell = get_cell_to_write(ws, row, col)
    existing_font = copy(cell.font) if cell.has_style else None
    existing_fill = copy(cell.fill) if cell.has_style else None
    existing_border = copy(cell.border) if cell.has_style else None
    existing_protection = copy(cell.protection) if cell.has_style else None
    cell.value = value
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right')
    if existing_font:
        cell.font = existing_font
    if existing_fill:
        cell.fill = existing_fill
    if existing_border:
        cell.border = existing_border
    if existing_protection:
        cell.protection = existing_protection

def safe_set_value(ws, row, col, value):
    if row <= 13:
        return
    cell = get_cell_to_write(ws, row, col)
    existing_font = copy(cell.font) if cell.has_style else None
    existing_fill = copy(cell.fill) if cell.has_style else None
    existing_border = copy(cell.border) if cell.has_style else None
    existing_protection = copy(cell.protection) if cell.has_style else None
    cell.value = value
    cell.alignment = Alignment(horizontal='right')
    if existing_font:
        cell.font = existing_font
    if existing_fill:
        cell.fill = existing_fill
    if existing_border:
        cell.border = existing_border
    if existing_protection:
        cell.protection = existing_protection

def align_numeric_cells(ws):
    logger.info("Выравнивание числовых ячеек по правому краю...")
    for row in range(14, ws.max_row + 1):
        for col in NUMERIC_COLUMNS:
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and not is_cell_merged(ws, row, col):
                existing_font = copy(cell.font) if cell.has_style else None
                existing_fill = copy(cell.fill) if cell.has_style else None
                existing_border = copy(cell.border) if cell.has_style else None
                existing_protection = copy(cell.protection) if cell.has_style else None
                cell.alignment = Alignment(horizontal='right')
                if existing_font:
                    cell.font = existing_font
                if existing_fill:
                    cell.fill = existing_fill
                if existing_border:
                    cell.border = existing_border
                if existing_protection:
                    cell.protection = existing_protection

def get_cell_value(ws, row, col):
    if is_cell_merged(ws, row, col):
        cell = get_cell_to_write(ws, row, col)
        return cell.value
    return ws.cell(row=row, column=col).value

def safe_set_top_table_value(ws, row, col, value):
    cell = get_cell_to_write(ws, row, col)
    cell.value = value
    cell.number_format = '#,##0.00' if col == 5 else '0.00%'
    cell.alignment = Alignment(horizontal='right')

def get_interval_col(days):
    if days <= 0:
        return COLUMNS['NOT_OVERDUE']
    elif 1 <= days <= 15:
        return COLUMNS['INTERVAL_1_15']
    elif 16 <= days <= 29:
        return COLUMNS['INTERVAL_16_29']
    elif 30 <= days <= 89:
        return COLUMNS['INTERVAL_30_89']
    elif 90 <= days <= 179:
        return COLUMNS['INTERVAL_90_179']
    else:
        return COLUMNS['INTERVAL_180_PLUS']

def clear_all_intervals(ws, row):
    if row <= 13:
        return
    interval_cols = [
        COLUMNS['NOT_OVERDUE'], COLUMNS['INTERVAL_1_15'],
        COLUMNS['INTERVAL_16_29'], COLUMNS['INTERVAL_30_89'],
        COLUMNS['INTERVAL_90_179'], COLUMNS['INTERVAL_180_PLUS'],
    ]
    for col in interval_cols:
        safe_set_number_format(ws, row, col, 0)

DOCUMENT_KEYWORDS = [
    'Акт', 'Реализация', 'Корректировка', 'Поступление',
    'Взаимозачет', 'Взаимозачёт', 'Списание', 'УПД', 'Счет-фактура',
    'Товарная накладная', 'ТОРГ-12', 'Универсальный передаточный'
]

def find_structure(ws):
    filials = []
    kontragents = []
    dogovors = []
    documents = []
    total_row = None
    for row in range(14, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if not cell_value:
            continue
        str_val = str(cell_value).strip()
        if str_val.startswith('ДТ '):
            filials.append(row)
        elif 'Итого' in str_val or 'ИТОГО' in str_val:
            total_row = row
        elif any(keyword in str_val for keyword in DOCUMENT_KEYWORDS):
            documents.append(row)
        elif str_val.startswith('Договор') or (str_val.startswith('договор') and not any(kw in str_val for kw in DOCUMENT_KEYWORDS)):
            dogovors.append(row)
        else:
            if len(str_val) > 2 and not str_val[0].isdigit():
                kontragents.append(row)
    return filials, kontragents, dogovors, documents, total_row

def get_cell_indent(ws, row, col=1):
    cell = ws.cell(row=row, column=col)
    if cell.alignment and cell.alignment.indent is not None:
        return int(cell.alignment.indent)
    return 0

def build_hierarchy(ws, start_row=14):
    nodes = {}
    stack = []
    total_row = None
    filial_rows = []
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        value = cell.value
        if value is None:
            continue
        str_val = str(value).strip()
        if not str_val:
            continue
        if 'Итого' in str_val or 'ИТОГО' in str_val:
            total_row = row
            continue
        indent = get_cell_indent(ws, row, 1)
        nodes[row] = {'indent': indent, 'children': [], 'parent': None}
        if indent == 0 and str_val.startswith('ДТ '):
            filial_rows.append(row)
        while stack and stack[-1][0] >= indent:
            stack.pop()
        if stack:
            parent_row = stack[-1][1]
            nodes[row]['parent'] = parent_row
            nodes[parent_row]['children'].append(row)
        stack.append((indent, row))
    return nodes, total_row, filial_rows

def recalc_totals(ws):
    logger.info("=== ПЕРЕСЧЁТ ИТОГОВ (по отступам, только листья) ===")
    nodes, total_row, filial_rows = build_hierarchy(ws)
    if not nodes:
        logger.warning("Не найдено строк с данными для пересчёта")
        return
    logger.info(f"Найдено узлов: {len(nodes)}, филиалов: {len(filial_rows)}, итого: строка {total_row}")
    leaves = set()
    for row, node in nodes.items():
        if not node['children']:
            leaves.add(row)
    logger.info(f"Листьев: {len(leaves)}")

    def get_all_leaves_under(row):
        node = nodes.get(row)
        if not node:
            return set()
        if row in leaves:
            return {row}
        result = set()
        for child in node['children']:
            result |= get_all_leaves_under(child)
        return result

    def sum_leaves(leaf_set, col):
        total = 0
        for r in leaf_set:
            val = get_cell_value(ws, r, col)
            if isinstance(val, (int, float)):
                total += val
        return total

    def max_days_in_leaf_set(leaf_set):
        max_val = 0
        for r in leaf_set:
            val = get_cell_value(ws, r, COLUMNS['DAYS'])
            if isinstance(val, (int, float)) and val > max_val:
                max_val = val
        return max_val

    sorted_rows = sorted(nodes.keys(), reverse=True)
    recalculated_count = 0
    for row in sorted_rows:
        node = nodes[row]
        if row in leaves:
            continue
        leaf_set = get_all_leaves_under(row)
        if not leaf_set:
            continue
        recalculated_count += 1
        for col in SUM_COLUMNS:
            total = sum_leaves(leaf_set, col)
            safe_set_number_format(ws, row, col, total)
        max_day = max_days_in_leaf_set(leaf_set)
        safe_set_value(ws, row, COLUMNS['DAYS'], max_day)
    logger.info(f"Пересчитано узлов: {recalculated_count}")

    if total_row and filial_rows:
        all_leaves_under_filials = set()
        for fil_row in filial_rows:
            all_leaves_under_filials |= get_all_leaves_under(fil_row)
        logger.info(f"Общий итог стр.{total_row}: {len(all_leaves_under_filials)} листьев под филиалами")
        for col in SUM_COLUMNS:
            total = sum_leaves(all_leaves_under_filials, col)
            safe_set_number_format(ws, total_row, col, total)
        max_day = max_days_in_leaf_set(all_leaves_under_filials)
        safe_set_value(ws, total_row, COLUMNS['DAYS'], max_day)
    logger.info("=== ПЕРЕСЧЁТ ИТОГОВ ЗАВЕРШЁН ===")

def update_top_table(ws, total_row):
    logger.info("=== ОБНОВЛЕНИЕ ВЕРХНЕЙ ТАБЛИЦЫ ===")
    if not total_row:
        logger.warning("Не найдена итоговая строка")
        return
    t_value = get_cell_value(ws, total_row, COLUMNS['NOT_OVERDUE']) or 0
    u_value = get_cell_value(ws, total_row, COLUMNS['INTERVAL_1_15']) or 0
    v_value = get_cell_value(ws, total_row, COLUMNS['INTERVAL_16_29']) or 0
    w_value = get_cell_value(ws, total_row, COLUMNS['INTERVAL_30_89']) or 0
    x_value = get_cell_value(ws, total_row, COLUMNS['INTERVAL_90_179']) or 0
    y_value = get_cell_value(ws, total_row, COLUMNS['INTERVAL_180_PLUS']) or 0
    l_value = get_cell_value(ws, total_row, COLUMNS['DEBT_AMOUNT']) or 0
    safe_set_top_table_value(ws, 2, 5, t_value)
    safe_set_top_table_value(ws, 2, 6, t_value / l_value if l_value else 0)
    safe_set_top_table_value(ws, 3, 5, u_value)
    safe_set_top_table_value(ws, 3, 6, u_value / l_value if l_value else 0)
    safe_set_top_table_value(ws, 4, 5, v_value)
    safe_set_top_table_value(ws, 4, 6, v_value / l_value if l_value else 0)
    safe_set_top_table_value(ws, 5, 5, w_value)
    safe_set_top_table_value(ws, 5, 6, w_value / l_value if l_value else 0)
    safe_set_top_table_value(ws, 6, 5, x_value)
    safe_set_top_table_value(ws, 6, 6, x_value / l_value if l_value else 0)
    safe_set_top_table_value(ws, 7, 5, y_value)
    safe_set_top_table_value(ws, 7, 6, y_value / l_value if l_value else 0)
    safe_set_top_table_value(ws, 8, 5, l_value)
    safe_set_top_table_value(ws, 8, 6, 1.0)
    logger.info("Верхняя таблица обновлена")

def copy_worksheet_full(ws, wb):
    logger.info(f"Копирование ячеек листа '{ws.title}'...")
    new_ws = wb.create_sheet(ws.title)
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)
                new_cell.protection = copy(cell.protection)
            if cell.data_type == 'f':
                new_cell.data_type = 'f'
    for merged_range in ws.merged_cells.ranges:
        new_ws.merge_cells(str(merged_range))
    for col_letter, col_dim in ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = col_dim.width
        if col_dim.outline_level:
            new_ws.column_dimensions[col_letter].outline_level = col_dim.outline_level
        if col_dim.hidden:
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    for row_num, row_dim in ws.row_dimensions.items():
        new_ws.row_dimensions[row_num].height = row_dim.height
        if row_dim.outline_level:
            new_ws.row_dimensions[row_num].outline_level = row_dim.outline_level
        if row_dim.hidden:
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden
    if hasattr(ws.sheet_properties, 'outlinePr') and ws.sheet_properties.outlinePr:
        new_ws.sheet_properties.outlinePr.summaryBelow = ws.sheet_properties.outlinePr.summaryBelow
        new_ws.sheet_properties.outlinePr.summaryRight = ws.sheet_properties.outlinePr.summaryRight
    if ws.print_options:
        for attr in ['grid_lines', 'grid_lines_set', 'horizontal_centered', 'vertical_centered']:
            if hasattr(ws.print_options, attr):
                setattr(new_ws.print_options, attr, getattr(ws.print_options, attr))
    if ws.page_setup:
        for attr in ['orientation', 'paperSize', 'scale', 'fitToHeight', 'fitToWidth',
                     'pageOrder', 'blackAndWhite', 'draft', 'cellComments', 'errors']:
            if hasattr(ws.page_setup, attr) and getattr(ws.page_setup, attr) is not None:
                setattr(new_ws.page_setup, attr, getattr(ws.page_setup, attr))
    if ws.page_margins:
        for attr in ['left', 'right', 'top', 'bottom', 'header', 'footer']:
            if hasattr(ws.page_margins, attr):
                setattr(ws.page_margins, attr, getattr(ws.page_margins, attr))
    if ws.sheet_view and ws.sheet_view.pane:
        pane = ws.sheet_view.pane
        try:
            new_ws.sheet_view.pane = Pane(
                xSplit=pane.xSplit, ySplit=pane.ySplit,
                topLeftCell=pane.topLeftCell, activePane=pane.activePane, state=pane.state
            )
        except (TypeError, AttributeError):
            logger.warning(f"Пропущена настройка freeze panes (несовместимость версий)")
    if ws.auto_filter.ref:
        new_ws.auto_filter.ref = ws.auto_filter.ref
    for cf_rule in ws.conditional_formatting._cf_rules:
        new_ws.conditional_formatting._cf_rules.append(cf_rule)
    if hasattr(ws, 'tables') and ws.tables:
        for table in ws.tables:
            new_ws.tables.add(table)
    logger.info(f"Лист '{ws.title}' скопирован с полным форматированием")
    return new_ws

# ============================================================
# API: СВЕРКА ДЕБИТОРСКОЙ ЗАДОЛЖЕННОСТИ
# ============================================================
@app.route('/save-excel', methods=['POST'])
@limiter.limit("5 per minute")  # Ограничение для тяжёлой операции
def save_excel():
    try:
        file = request.files['file']
        data = json.loads(request.form['data'])

        logger.info(f"=== ПОЛУЧЕН ЗАПРОС ===")
        logger.info(f"Файл: {file.filename}")
        logger.info(f"Документов для обновления: {len(data['updatedDocuments'])}")

        wb = openpyxl.load_workbook(io.BytesIO(file.read()))

        logger.info("=== ШАГ 0: ПЕРЕИМЕНОВАНИЕ ЛИСТОВ ===")
        all_sheets = list(wb.sheetnames)
        logger.info(f"Исходные листы: {all_sheets}")

        siuat_sheets = []
        summary_sheets = []
        main_sheets = []

        for sheet_name in all_sheets:
            if sheet_name.startswith('Свод ДЗ СИ УАТ') or sheet_name.startswith('Свод ДЗ СИ'):
                siuat_sheets.append(sheet_name)
            elif 'Сводные' in sheet_name or 'Сводная' in sheet_name:
                summary_sheets.append(sheet_name)
            else:
                test_ws = wb[sheet_name]
                is_main = False
                for r in range(1, min(30, test_ws.max_row + 1)):
                    cell_val = test_ws.cell(row=r, column=1).value
                    if cell_val and str(cell_val).strip().startswith('ДТ '):
                        is_main = True
                        break
                if is_main:
                    main_sheets.append(sheet_name)

        logger.info(f"main_sheets={main_sheets}, siuat_sheets={siuat_sheets}, summary_sheets={summary_sheets}")

        temp_names = {}
        temp_counter = 0

        for sheet_name in all_sheets:
            if sheet_name.startswith('Свод ДЗ СИ УАТ') or sheet_name.startswith('Свод ДЗ СИ'):
                temp_counter += 1
                temp_names[sheet_name] = f'__temp_siuat_{temp_counter}__'
            elif 'Сводные' in sheet_name or 'Сводная' in sheet_name:
                temp_counter += 1
                temp_names[sheet_name] = f'__temp_summary_{temp_counter}__'

        for idx, sheet_name in enumerate(main_sheets):
            if idx == 0:
                temp_names[sheet_name] = '__temp_main__'
            elif idx == 1:
                temp_names[sheet_name] = '__temp_siuat_from_main__'
            else:
                temp_names[sheet_name] = f'__temp_extra_{idx}__'

        for old_name, temp_name in temp_names.items():
            if old_name != temp_name and old_name in wb.sheetnames:
                wb[old_name].title = temp_name
                logger.info(f"Шаг 1: '{old_name}' → '{temp_name}'")

        for sn in list(wb.sheetnames):
            if sn == '__temp_main__':
                wb[sn].title = 'Свод ДЗ'
                logger.info("Шаг 2: '__temp_main__' → 'Свод ДЗ'")
            elif sn == '__temp_siuat_from_main__':
                wb[sn].title = 'Свод ДЗ СИ УАТ'
                logger.info("Шаг 2: '__temp_siuat_from_main__' → 'Свод ДЗ СИ УАТ'")
            elif sn.startswith('__temp_siuat_'):
                wb[sn].title = 'Свод ДЗ СИ УАТ'
                logger.info(f"Шаг 2: '{sn}' → 'Свод ДЗ СИ УАТ'")
            elif sn.startswith('__temp_summary_'):
                wb[sn].title = 'Сводные таблицы'
                logger.info(f"Шаг 2: '{sn}' → 'Сводные таблицы'")
            elif sn.startswith('__temp_extra_'):
                del wb[sn]
                logger.info(f"Удалён лишний лист: {sn}")

        logger.info(f"Листы после переименования: {wb.sheetnames}")
        logger.info("=== ПЕРЕИМЕНОВАНИЕ ЗАВЕРШЕНО ===")

        ws = wb['Свод ДЗ']
        wb._active_sheet_index = wb.sheetnames.index('Свод ДЗ')

        today = datetime.now().date()
        logger.info(f"Текущая дата: {today}")

        _, _, _, _, total_row = find_structure(ws)

        updated_rows = set()
        docs_with_date = 0
        docs_without_date = 0
        docs_overdue = 0
        docs_not_overdue = 0

        for idx, item in enumerate(data['updatedDocuments']):
            row_number = item['rowNumber']
            debt_amount = float(item['amount'])
            expected_date_str = item['date']
            doc_name = item.get('documentName', 'неизвестно')
            action = item.get('action', '')

            expected_date = None
            if expected_date_str and expected_date_str not in ('null', 'None', ''):
                try:
                    clean_str = expected_date_str.strip().strip('"').strip("'")
                    if clean_str:
                        expected_date = datetime.fromisoformat(clean_str.replace('Z', '+00:00')).date()
                except Exception as e:
                    logger.warning(f"Ошибка парсинга даты: {e}")
                    expected_date = None

            if idx < 10 or expected_date is None:
                status = "БЕЗ ДАТЫ" if expected_date is None else f"дата: {expected_date}"
                logger.info(f"[{idx+1}] Строка {row_number}: '{doc_name[:50]}...' — {status}, сумма: {debt_amount:,.2f}")

            clear_all_intervals(ws, row_number)

            if expected_date is None or expected_date >= today:
                reason = 'нет даты в файле поступлений' if expected_date is None else 'дата в будущем или сегодня'
                logger.info(f"  → НЕ ПРОСРОЧЕНО (причина: {reason})")
                safe_set_number_format(ws, row_number, COLUMNS['OVERDUE'], 0)
                safe_set_value(ws, row_number, COLUMNS['DAYS'], 0)
                safe_set_number_format(ws, row_number, COLUMNS['NOT_OVERDUE'], debt_amount)
                updated_rows.add(row_number)
                docs_not_overdue += 1
                if expected_date is None:
                    docs_without_date += 1
                else:
                    docs_with_date += 1
            elif expected_date < today:
                days_overdue = (today - expected_date).days
                interval_col = get_interval_col(days_overdue)
                logger.info(f"  → ПРОСРОЧЕНО на {days_overdue} дн. (интервал: {interval_col})")
                safe_set_number_format(ws, row_number, COLUMNS['OVERDUE'], debt_amount)
                safe_set_value(ws, row_number, COLUMNS['DAYS'], days_overdue)
                safe_set_number_format(ws, row_number, interval_col, debt_amount)
                updated_rows.add(row_number)
                docs_overdue += 1
                docs_with_date += 1

        if updated_rows:
            logger.info(f"=== СТАТИСТИКА ОБРАБОТКИ ===")
            logger.info(f"Обновлено строк: {len(updated_rows)}")
            logger.info(f"  Документов с датой: {docs_with_date}")
            logger.info(f"  Документов БЕЗ даты: {docs_without_date}")
            logger.info(f"  Просрочено: {docs_overdue}")
            logger.info(f"  Не просрочено: {docs_not_overdue}")
            logger.info("============================")

            recalc_totals(ws)
            if total_row:
                update_top_table(ws, total_row)
        else:
            logger.info("Нет обновлённых строк")

        align_numeric_cells(ws)

        current_day_data_from_file = extract_filial_overdue(ws)
        data['currentDayData'] = current_day_data_from_file
        total_debt_data = extract_total_row_debt(ws, total_row)
        logger.info(f"Из итоговой строки ДТ: общая ДЗ={total_debt_data['totalDebt']}, ПДЗ={total_debt_data['totalOverdue']}")

        siuat_file = request.files.get('siUatFile')
        siuat_total_debt = 0
        siuat_total_overdue = 0
        siuat_sheet_created = False

        logger.info(f"=== ОТЛАДКА СИ УАТ ===")
        logger.info(f"siuat_file: {siuat_file}")
        logger.info(f"siuat_file.filename: {siuat_file.filename if siuat_file else 'None'}")

        summary_siuat = data.get('summarySIUAT', {})
        if summary_siuat:
            siuat_total_debt = summary_siuat.get('totalDebt', 0) or 0
            siuat_total_overdue = summary_siuat.get('totalOverdue', 0) or 0
            logger.info(f"Получены данные СИ УАТ с фронтенда: общая ДЗ={siuat_total_debt}, ПДЗ={siuat_total_overdue}")

        if siuat_file and siuat_file.filename:
            logger.info(f"=== ДОБАВЛЯЕМ ЛИСТ 'Свод ДЗ СИ УАТ' из файла {siuat_file.filename} ===")
            try:
                sheets_to_delete = [sn for sn in wb.sheetnames if sn.startswith('Свод ДЗ СИ УАТ')]
                for sheet_name in sheets_to_delete:
                    del wb[sheet_name]
                    logger.info(f"Удалён лист: {sheet_name}")

                siuat_file_content = siuat_file.read()
                logger.info(f"Размер файла СИ УАТ: {len(siuat_file_content)} байт")
                siuat_wb = openpyxl.load_workbook(io.BytesIO(siuat_file_content))
                logger.info(f"Листы в файле СИ УАТ: {siuat_wb.sheetnames}")

                siuat_ws = siuat_wb.worksheets[0]
                logger.info(f"Используем лист: '{siuat_ws.title}', размер: {siuat_ws.max_row} строк")

                new_siuat_ws = copy_worksheet_full(siuat_ws, wb)
                new_siuat_ws.title = 'Свод ДЗ СИ УАТ'
                siuat_sheet_created = True
                logger.info(f"Лист скопирован и переименован в 'Свод ДЗ СИ УАТ'")

                # Закрываем промежуточный workbook для освобождения памяти
                siuat_wb.close()
                del siuat_wb
                gc.collect()

                if siuat_total_debt == 0:
                    logger.info("Вызываем extract_siuat_totals_by_max...")
                    file_debt, file_overdue = extract_siuat_totals_by_max(new_siuat_ws)
                    if file_debt > 0:
                        siuat_total_debt = file_debt
                        siuat_total_overdue = file_overdue
                        logger.info(f"Используем данные из файла: общая ДЗ={siuat_total_debt}, ПДЗ={siuat_total_overdue}")
                    else:
                        logger.warning("ВНИМАНИЕ: extract_siuat_totals_by_max вернул 0!")
                else:
                    logger.info("Используем данные с фронтенда, файл не сканируем")
            except Exception as e:
                logger.error(f"!!! Ошибка при добавлении листа СИ УАТ: {e}")
                traceback.print_exc()
        else:
            logger.info("Файл СИ УАТ не загружен (siuat_file=None или filename пустой)")

        logger.info("=== СОЗДАЁМ ЛИСТ 'Сводные таблицы' ===")
        try:
            if 'Сводные таблицы' in wb.sheetnames:
                del wb['Сводные таблицы']
                logger.info("Удалён существующий лист 'Сводные таблицы'")
            summary_ws = wb.create_sheet('Сводные таблицы')
            create_summary_sheet(
                summary_ws, data,
                total_debt=total_debt_data['totalDebt'],
                total_overdue=total_debt_data['totalOverdue'],
                siuat_total_debt=siuat_total_debt,
                siuat_total_overdue=siuat_total_overdue,
            )
            logger.info("Лист 'Сводные таблицы' создан")
        except Exception as e:
            logger.error(f"!!! Ошибка создания сводных таблиц: {e}")
            traceback.print_exc()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Закрываем workbook для освобождения памяти
        wb.close()
        del wb
        gc.collect()

        logger.info("=== ФАЙЛ УСПЕШНО ОБРАБОТАН, ОТПРАВЛЯЕМ ===")

        filial_data_json = json.dumps(current_day_data_from_file, ensure_ascii=False)
        filial_data_encoded = base64.b64encode(filial_data_json.encode("utf-8")).decode("ascii")

        response = send_file(
            output,
            as_attachment=True,
            download_name=f"ДЗ_обновленный_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Filial-Data"] = filial_data_encoded

        return response

    except Exception as e:
        logger.error(f"!!! ОШИБКА в save_excel: {e}")
        traceback.print_exc()
        return {'error': str(e)}, 500

# ============================================================
# API: ДАННЫЕ ПРЕДЫДУЩЕГО ДНЯ
# ============================================================
@app.route('/api/previous-day-data', methods=['POST', 'GET'])
def api_previous_day_data():
    try:
        if request.method == 'POST':
            data = request.get_json()
            if not data:
                return jsonify({'error': 'Нет данных'}), 400

            swipe_date = data.get('date')
            filial_data = data.get('data', {})
            summary_dt = data.get('summaryDT', {})
            summary_siuat = data.get('summarySIUAT', {})

            if not swipe_date:
                return jsonify({'error': 'Не указана дата'}), 400

            logger.info(f"=== СОХРАНЕНИЕ ДАННЫХ ПРЕДЫДУЩЕГО ДНЯ В БД ===")
            logger.info(f"Дата: {swipe_date}, Филиалов: {len(filial_data)}")

            result = db.save_previous_day_data(swipe_date, filial_data, summary_dt, summary_siuat)
            return jsonify(result)

        else:  # GET
            requested_date = request.args.get('date')
            if not requested_date:
                return jsonify({'error': 'Не указана дата (параметр ?date=YYYY-MM-DD)'}), 400

            logger.info(f"=== ЗАГРУЗКА ДАННЫХ ПРЕДЫДУЩЕГО ДНЯ ИЗ БД ===")
            logger.info(f"Запрошенная дата: {requested_date}")

            filial_data = db.get_previous_day_data(requested_date)
            summary_dt = {}
            summary_siuat = {}
            actual_date = requested_date

            if filial_data:
                conn = None
                try:
                    conn = db.get_connection()
                    cur = conn.cursor()
                    cur.execute("""
                        SELECT DISTINCT swipe_date
                        FROM swipe_history
                        WHERE swipe_date = %s
                    """, (requested_date,))
                    result = cur.fetchone()

                    if not result:
                        last_date = db.get_last_available_swipe_date(requested_date)
                        if last_date:
                            actual_date = last_date
                            logger.info(f"🔄 Данные найдены за предыдущую дату: {actual_date}")

                    cur.close()
                except Exception as e:
                    logger.warning(f"⚠️ Не удалось определить фактическую дату: {e}")
                finally:
                    if conn:
                        db.return_connection(conn)

                summary_result = db.get_summary_data(actual_date)
                if summary_result:
                    summary_dt = summary_result.get('summaryDT', {})
                    summary_siuat = summary_result.get('summarySIUAT', {})
                    logger.info(f"✅ Загружены сводные данные за {actual_date}")

                logger.info(f"✅ Загружено {len(filial_data)} филиалов за {actual_date}")
                return jsonify({
                    'success': True,
                    'data': filial_data,
                    'date': actual_date,
                    'requested_date': requested_date,
                    'summaryDT': summary_dt,
                    'summarySIUAT': summary_siuat
                })
            else:
                logger.warning(f"⚠️ Данные не найдены ни за {requested_date}, ни за предыдущие даты")
                return jsonify({
                    'success': True,
                    'data': {},
                    'date': requested_date,
                    'requested_date': requested_date,
                    'message': 'Данные не найдены',
                    'summaryDT': {},
                    'summarySIUAT': {}
                })

    except Exception as e:
        logger.error(f"❌ Ошибка API previous-day-data: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ============================================================
# API: РАСПРЕДЕЛЕНИЕ НАЛОГОВ ПО ПОДРАЗДЕЛЕНИЯМ
# ============================================================
@app.route('/api/allocate-cashflow', methods=['POST'])
@limiter.limit("3 per minute")  # Тяжёлая операция
def api_allocate_cashflow():
    try:
        logger.info("=== ЗАПРОС НА РАСПРЕДЕЛЕНИЕ НАЛОГОВ ===")

        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'Файл не загружен'}), 400

        settings_str = request.form.get('settings', '{}')
        settings = json.loads(settings_str)

        logger.info(f"Файл: {file.filename}, Настройки: {settings}")

        file_bytes = file.read()
        logger.info(f"Размер файла: {len(file_bytes)} байт")

        result = allocator.allocate_cashflow(
            file_bytes=file_bytes,
            strategy=settings.get('strategy', 'revenue'),
            rounding=settings.get('rounding', 'kopeks'),
            period_type=settings.get('periodType', 'month'),
            period_value=settings.get('periodValue', ''),
            date_from=settings.get('dateFrom', ''),
            date_to=settings.get('dateTo', ''),
            exclude_articles=settings.get('excludeArticles', [])
        )

        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Ошибка распределения')}), 500

        summary = {
            'success': True,
            'filials_count': result.get('filials_count', 0),
            'total_revenue': result.get('total_revenue', 0),
            'total_taxes': result.get('total_taxes', 0),
            'shares': result.get('shares', {}),
            'period': settings.get('periodType', 'month')
        }

        summary_json = json.dumps(summary, ensure_ascii=False)
        summary_encoded = base64.b64encode(summary_json.encode("utf-8")).decode("ascii")

        logger.info(f"✅ Распределение завершено: {result.get('filials_count')} подразделений")

        response = send_file(
            result['file'],
            as_attachment=True,
            download_name=f"ДДС_налоги_распределены_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response.headers["X-Allocation-Summary"] = summary_encoded

        return response

    except Exception as e:
        logger.error(f"!!! ОШИБКА РАСПРЕДЕЛЕНИЯ НАЛОГОВ: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ СВЕРКИ ДЗ
# ============================================================
def create_summary_sheet(ws, data, total_debt=0, total_overdue=0, siuat_total_debt=0, siuat_total_overdue=0):
    logger.info("Создание листа 'Сводные таблицы'...")

    current_date = data.get('currentDate', datetime.now().strftime('%Y-%m-%d'))
    previous_date = data.get('previousDate', '')
    current_day_data = data.get('currentDayData', {})
    previous_day_data = data.get('previousDayData', {})

    title_font = Font(bold=True, size=14)
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    number_format = '#,##0.00'
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    row = 1
    ws.cell(row=row, column=1, value='Динамика по подразделениям').font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    def format_date_ddmmyyyy(date_str):
        if not date_str or date_str in ('предыдущий рабочий день',):
            return date_str
        try:
            dt = datetime.fromisoformat(date_str)
            return dt.strftime('%d.%m.%Y')
        except (ValueError, TypeError):
            return date_str

    current_date_formatted = format_date_ddmmyyyy(current_date)
    previous_date_formatted = format_date_ddmmyyyy(previous_date)

    headers = ['Подразделение', current_date_formatted, previous_date_formatted, 'Динамика']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    row += 1

    all_filials = sorted(set(list(current_day_data.keys()) + list(previous_day_data.keys())))
    total_current = total_previous = total_delta = 0

    for filial in all_filials:
        cv = current_day_data.get(filial, 0)
        pv = previous_day_data.get(filial, 0)
        d = pv - cv
        total_current += cv
        total_previous += pv
        total_delta += d
        ws.cell(row=row, column=1, value=filial).border = thin_border
        cc = ws.cell(row=row, column=2, value=cv)
        cc.number_format = number_format
        cc.border = thin_border
        cc.alignment = Alignment(horizontal='right')
        cp = ws.cell(row=row, column=3, value=pv)
        cp.number_format = number_format
        cp.border = thin_border
        cp.alignment = Alignment(horizontal='right')
        cd = ws.cell(row=row, column=4, value=d)
        cd.number_format = number_format
        cd.border = thin_border
        cd.alignment = Alignment(horizontal='right')
        if d > 0:
            cd.fill = green_fill
        elif d < 0:
            cd.fill = red_fill
        row += 1

    ws.cell(row=row, column=1, value='Общий итог').font = total_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).border = thin_border
    for col, val in enumerate([total_current, total_previous, total_delta], 2):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = number_format
        c.font = total_font
        c.fill = total_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='right')
    row += 3

    ws.cell(row=row, column=1, value='Свод задолженности ДТ').font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2
    summary_dt = data.get('summaryDT', {})
    rows_dt = [
        ('общая ДЗ', total_debt),
        ('из них ПДЗ', total_overdue),
        ('в т.ч. Судебная', summary_dt.get('legal', 0)),
        ('не подлежащая к взысканию', summary_dt.get('notRecoverable', 0)),
        ('подлежащая к взысканию', summary_dt.get('recoverable', 0)),
    ]
    for label, value in rows_dt:
        cl = ws.cell(row=row, column=1, value=label)
        cl.border = thin_border
        if 'ПДЗ' in label:
            cl.font = Font(bold=True, color='FF0000')
        elif 'Судебная' in label:
            cl.font = Font(bold=True, color='0000FF')
        else:
            cl.font = Font(bold=True)
        cv = ws.cell(row=row, column=2, value=value)
        cv.number_format = number_format
        cv.border = thin_border
        cv.alignment = Alignment(horizontal='right')
        if 'ПДЗ' in label:
            cv.font = Font(bold=True, color='FF0000')
        elif 'Судебная' in label:
            cv.font = Font(bold=True, color='0000FF')
        else:
            cv.font = Font(bold=True)
        row += 1
    row += 3

    ws.cell(row=row, column=1, value='Свод задолженности СИ УАТ').font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2
    for label, value in [('Общая ДЗ', siuat_total_debt), ('Из них ПДЗ', siuat_total_overdue)]:
        cl = ws.cell(row=row, column=1, value=label)
        cl.border = thin_border
        cl.font = Font(bold=True)
        cv = ws.cell(row=row, column=2, value=value)
        cv.number_format = number_format
        cv.border = thin_border
        cv.alignment = Alignment(horizontal='right')
        if 'ПДЗ' in label:
            cv.font = Font(bold=True, color='FF0000')
        else:
            cv.font = Font(bold=True)
        row += 1
    logger.info("Лист 'Сводные таблицы' создан")


def extract_filial_overdue(ws):
    data = {}
    for r in range(14, ws.max_row + 1):
        v = get_cell_value(ws, r, 1)
        if v and str(v).strip().startswith('ДТ '):
            cell = ws.cell(row=r, column=1)
            indent = int(cell.alignment.indent) if cell.alignment and cell.alignment.indent is not None else 0
            if indent != 0:
                continue
            ov = get_cell_value(ws, r, COLUMNS['OVERDUE'])
            data[str(v).strip()] = ov if isinstance(ov, (int, float)) else 0
    for k in data:
        data[k] = round(data[k], 2)
    logger.info(f"=== ДАННЫЕ ФИЛИАЛОВ ===")
    for f, a in sorted(data.items()):
        logger.info(f"  {f}: {a:,.2f}")
    return data


def find_siuat_columns(ws):
    total_col = None
    overdue_col = None
    total_candidates = []
    overdue_candidates = []

    for r in range(1, min(21, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r, column=c).value
            if not cell_val:
                continue
            cell_str = str(cell_val).lower().strip()
            cell_str = ' '.join(cell_str.split())

            if 'всего' in cell_str:
                total_candidates.append((c, 0, cell_str))
            elif 'общая дз' in cell_str:
                total_candidates.append((c, 1, cell_str))
            elif 'общая задолженность' in cell_str:
                total_candidates.append((c, 2, cell_str))
            elif 'общая' in cell_str or 'total' in cell_str:
                total_candidates.append((c, 3, cell_str))

            if 'просроченно' in cell_str:
                overdue_candidates.append((c, 0, cell_str))
            elif 'просрочка' in cell_str:
                overdue_candidates.append((c, 1, cell_str))
            elif 'пдз' in cell_str:
                overdue_candidates.append((c, 2, cell_str))
            elif 'просроченная дз' in cell_str or 'просроченная задолженность' in cell_str:
                overdue_candidates.append((c, 3, cell_str))
            elif 'overdue' in cell_str:
                overdue_candidates.append((c, 3, cell_str))

    if total_candidates:
        total_candidates.sort(key=lambda x: (x[1], x[0]))
        best = total_candidates[0]
        total_col = best[0]
        logger.info(f"Найдена колонка 'всего': колонка {total_col} (приоритет {best[1]}, '{best[2]}')")

    if overdue_candidates:
        overdue_candidates.sort(key=lambda x: (x[1], x[0]))
        best = overdue_candidates[0]
        overdue_col = best[0]
        logger.info(f"Найдена колонка 'просроченно': колонка {overdue_col} (приоритет {best[1]}, '{best[2]}')")

    if total_col is None:
        total_col = 12
        logger.info(f"Fallback: колонка 'всего' = {total_col}")
    if overdue_col is None:
        overdue_col = 15
        logger.info(f"Fallback: колонка 'просроченно' = {overdue_col}")

    return total_col, overdue_col


def _parse_cell_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        cleaned = value.replace(' ', '').replace(',', '.')
        try:
            return float(cleaned)
        except ValueError:
            return 0
    return 0


def extract_siuat_totals_by_max(ws):
    total_col = 12
    overdue_col = 15
    logger.info(f"=== extract_siuat_totals_by_max START ===")
    logger.info(f"Лист: '{ws.title}', строк: {ws.max_row}, колонок: {ws.max_column}")

    total_debt = 0
    total_overdue = 0
    dt_count = 0

    for r in range(1, ws.max_row + 1):
        cell_val = get_cell_value(ws, r, 1)
        if not cell_val:
            continue
        str_val = str(cell_val).strip()
        if str_val.startswith('ДТ '):
            v_total = _parse_cell_number(get_cell_value(ws, r, total_col))
            v_overdue = _parse_cell_number(get_cell_value(ws, r, overdue_col))
            total_debt += v_total
            total_overdue += v_overdue
            dt_count += 1

    total_debt = round(total_debt, 2)
    total_overdue = round(total_overdue, 2)
    logger.info(f"Найдено подразделений ДТ: {dt_count}")
    logger.info(f"Результат: общая ДЗ={total_debt}, ПДЗ={total_overdue}")
    return total_debt, total_overdue


def extract_siuat_totals(ws):
    total_col, overdue_col = find_siuat_columns(ws)
    total_debt = 0
    total_overdue = 0

    for r in range(1, ws.max_row + 1):
        cell_val = get_cell_value(ws, r, 1)
        if cell_val:
            str_val = str(cell_val).strip().lower()
            if 'итог' in str_val or 'total' in str_val or 'всего' in str_val:
                v_total = get_cell_value(ws, r, total_col)
                v_overdue = get_cell_value(ws, r, overdue_col)
                total_debt = round(v_total, 2) if isinstance(v_total, (int, float)) else 0
                total_overdue = round(v_overdue, 2) if isinstance(v_overdue, (int, float)) else 0
                logger.info(f"СИ УАТ из строки '{cell_val}' (row {r}): общая ДЗ={total_debt}, ПДЗ={total_overdue}")
                break

    if total_debt == 0:
        logger.info("Строка 'ИТОГО' не найдена, используем fallback...")
        for r in range(ws.max_row, 0, -1):
            val = get_cell_value(ws, r, total_col)
            if isinstance(val, (int, float)) and val > 0:
                total_debt = round(val, 2)
                v_overdue = get_cell_value(ws, r, overdue_col)
                total_overdue = round(v_overdue, 2) if isinstance(v_overdue, (int, float)) else 0
                logger.info(f"СИ УАТ fallback строка {r}: общая ДЗ={total_debt}, ПДЗ={total_overdue}")
                break

    return total_debt, total_overdue


def extract_total_row_debt(ws, total_row):
    result = {'totalDebt': 0, 'totalOverdue': 0}
    if not total_row:
        return result
    td = get_cell_value(ws, total_row, COLUMNS['DEBT_AMOUNT'])
    to = get_cell_value(ws, total_row, COLUMNS['OVERDUE'])
    result['totalDebt'] = round(td, 2) if isinstance(td, (int, float)) else 0
    result['totalOverdue'] = round(to, 2) if isinstance(to, (int, float)) else 0
    return result


@app.route('/save-suppliers', methods=['POST'])
@limiter.limit("5 per minute")
def save_suppliers():
    try:
        file = request.files['file']
        data = json.loads(request.form['data'])

        logger.info(f"=== ПОЛУЧЕН ЗАПРОС НА ОБРАБОТКУ ОПЛАТ ПОСТАВЩИКАМ ===")
        logger.info(f"Файл: {file.filename}")
        logger.info(f"Сводных таблиц: {len(data.get('pivotTables', []))}")

        wb = openpyxl.load_workbook(io.BytesIO(file.read()))

        for pivot_table in data.get('pivotTables', []):
            sheet_name = pivot_table['sheetName']
            headers = pivot_table['headers']
            rows_data = pivot_table['data']

            logger.info(f"Обработка сводной таблицы: {sheet_name}")
            logger.info(f"  Подразделений: {len(headers)}")
            logger.info(f"  Контрагентов: {len(rows_data)}")

            if sheet_name not in wb.sheetnames:
                logger.warning(f"  Предупреждение: лист '{sheet_name}' не найден, создаём новый")
                ws = wb.create_sheet(sheet_name)
                last_row = 0
            else:
                ws = wb[sheet_name]
                last_row = ws.max_row
                logger.info(f"  Реестр заканчивается на строке: {last_row}")

            pivot_start_row = last_row + 4
            create_pivot_sheet_at_row(ws, headers, rows_data, 'Сводная таблица', pivot_start_row)
            logger.info(f"  Сводная таблица добавлена со строки: {pivot_start_row}")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Очистка памяти
        wb.close()
        del wb
        gc.collect()

        logger.info("=== ФАЙЛ ОПЛАТ УСПЕШНО ОБРАБОТАН, ОТПРАВЛЯЕМ ===")

        return send_file(
            output,
            as_attachment=True,
            download_name=f'Оплаты_поставщикам_{datetime.now().strftime("%Y-%m-%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"!!! ОШИБКА ПРИ ОБРАБОТКЕ ОПЛАТ: {e}")
        traceback.print_exc()
        return {'error': str(e)}, 500


def create_pivot_sheet_at_row(ws, headers, rows_data, title, start_row):
    logger.info(f"Создание сводной таблицы '{title}', начиная со строки {start_row}...")

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    explanation_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    number_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = start_row
    ws.cell(row=row, column=1, value='Сводная таблица оплат по подразделениям').font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers) + 3)
    row += 2

    header_cells = ['Контрагент'] + headers + ['Итого', 'Пояснение']
    for col, header in enumerate(header_cells, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    row += 1

    total_all = 0
    for item in rows_data:
        ws.cell(row=row, column=1, value=item['contractor']).border = thin_border
        total_sum = 0
        for col_idx, h in enumerate(headers, 2):
            value = item.get(h, 0)
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.number_format = number_format
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            total_sum += value
        total_all += total_sum
        cell_total = ws.cell(row=row, column=len(headers) + 2, value=total_sum)
        cell_total.number_format = number_format
        cell_total.border = thin_border
        cell_total.alignment = Alignment(horizontal='right')
        cell_explanation = ws.cell(row=row, column=len(headers) + 3, value=item.get('explanation', ''))
        cell_explanation.border = thin_border
        if item.get('explanation'):
            for col_idx in range(1, len(headers) + 4):
                ws.cell(row=row, column=col_idx).fill = explanation_fill
        row += 1

    ws.cell(row=row, column=1, value='ИТОГО').font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).border = thin_border

    for col_idx, h in enumerate(headers, 2):
        subtotal = sum(item.get(h, 0) for item in rows_data)
        cell = ws.cell(row=row, column=col_idx, value=subtotal)
        cell.number_format = number_format
        cell.font = Font(bold=True, size=11)
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')

    cell_grand_total = ws.cell(row=row, column=len(headers) + 2, value=total_all)
    cell_grand_total.number_format = number_format
    cell_grand_total.font = Font(bold=True, size=11)
    cell_grand_total.fill = total_fill
    cell_grand_total.border = thin_border
    cell_grand_total.alignment = Alignment(horizontal='right')

    ws.cell(row=row, column=len(headers) + 3).fill = total_fill
    ws.cell(row=row, column=len(headers) + 3).border = thin_border

    logger.info(f"Сводная таблица '{title}' создана, строк: {row - start_row + 1}")


# ============================================================
# API для отчётов (Dashboard)
# ============================================================

@app.route('/api/save-swipe-data', methods=['POST'])
def api_save_swipe_data():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400

        swipe_date = data.get('swipeDate')
        filial_data = data.get('filialData', {})
        counterparty_data = data.get('counterpartyData', {})
        total_debt = data.get('totalDebt', 0)
        total_overdue = data.get('totalOverdue', 0)
        summary_dt = data.get('summaryDT', {})
        summary_siuat = data.get('summarySIUAT', {})

        if not swipe_date:
            return jsonify({'error': 'Не указана дата сверки'}), 400

        logger.info(f"=== СОХРАНЕНИЕ ДАННЫХ СВЕРКИ В БД ===")
        logger.info(f"Дата: {swipe_date}, Филиалов: {len(filial_data)}, Контрагентов: {len(counterparty_data)}")

        result = db.save_swipe_data(
            swipe_date, filial_data, counterparty_data,
            total_debt, total_overdue, summary_dt, summary_siuat
        )
        return jsonify(result)

    except Exception as e:
        logger.error(f"❌ Ошибка API save-swipe-data: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/swipe-dates', methods=['GET'])
def api_swipe_dates():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        dates = db.get_swipe_dates(from_date, to_date)
        return jsonify({'success': True, 'data': dates})
    except Exception as e:
        logger.error(f"❌ Ошибка API swipe-dates: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/swipe-raw', methods=['GET'])
def api_swipe_raw():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        filial = request.args.get('filial')
        if not from_date or not to_date:
            return jsonify({'error': 'Укажите период (from и to)'}), 400
        result = db.get_swipe_raw_data(from_date, to_date, filial)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API swipe-raw: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/filial-trend', methods=['GET'])
def api_filial_trend():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        filial = request.args.get('filial')
        if not from_date or not to_date:
            return jsonify({'error': 'Укажите период (from и to)'}), 400
        result = db.get_filial_trend(from_date, to_date, filial)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API filial-trend: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/counterparty-trend', methods=['GET'])
def api_counterparty_trend():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        filial = request.args.get('filial')
        counterparty = request.args.get('counterparty')
        if not from_date or not to_date:
            return jsonify({'error': 'Укажите период (from и to)'}), 400
        result = db.get_counterparty_trend(from_date, to_date, filial, counterparty)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API counterparty-trend: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/filial-list', methods=['GET'])
def api_filial_list():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        result = db.get_filial_list(from_date, to_date)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API filial-list: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/counterparty-list', methods=['GET'])
def api_counterparty_list():
    try:
        filial = request.args.get('filial')
        result = db.get_counterparty_list(filial)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API counterparty-list: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/summary', methods=['GET'])
def api_summary():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        if not from_date or not to_date:
            return jsonify({'error': 'Укажите период (from и to)'}), 400
        result = db.get_summary(from_date, to_date)
        if result:
            return jsonify({'success': True, 'data': result})
        else:
            return jsonify({'success': True, 'data': None, 'message': 'Нет данных за указанный период'})
    except Exception as e:
        logger.error(f"❌ Ошибка API summary: {e}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# API: НАСТРОЙКИ ПАРСЕРА — МАППИНГ СЧЕТОВ
# ============================================================

@app.route('/api/account-mapping', methods=['GET'])
def api_get_account_mapping():
    try:
        all_records = request.args.get('all', 'false').lower() == 'true'
        result = db.get_account_mapping(active_only=not all_records)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API account-mapping GET: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/account-mapping', methods=['POST'])
def api_add_account_mapping():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400
        account_number = data.get('account_number', '').strip()
        company_name = data.get('company_name', '').strip()
        bank_name = data.get('bank_name', '').strip()
        if not account_number or not company_name:
            return jsonify({'error': 'Номер счёта и название компании обязательны'}), 400
        result = db.add_account_mapping(account_number, company_name, bank_name)
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify({'error': result.get('error', 'Ошибка сохранения')}), 500
    except Exception as e:
        logger.error(f"❌ Ошибка API account-mapping POST: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/account-mapping/<int:mapping_id>', methods=['PUT'])
def api_update_account_mapping(mapping_id):
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400
        result = db.update_account_mapping(
            mapping_id,
            account_number=data.get('account_number'),
            company_name=data.get('company_name'),
            bank_name=data.get('bank_name'),
            is_active=data.get('is_active')
        )
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify({'error': result.get('error', 'Ошибка обновления')}), 500
    except Exception as e:
        logger.error(f"❌ Ошибка API account-mapping PUT: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/account-mapping/<int:mapping_id>', methods=['DELETE'])
def api_delete_account_mapping(mapping_id):
    try:
        result = db.delete_account_mapping(mapping_id)
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify({'error': result.get('error', 'Ошибка удаления')}), 500
    except Exception as e:
        logger.error(f"❌ Ошибка API account-mapping DELETE: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/companies', methods=['GET'])
def api_companies_list():
    try:
        result = db.get_companies_list()
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API companies: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/banks', methods=['GET'])
def api_banks_list():
    try:
        result = db.get_banks_list()
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"❌ Ошибка API banks: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/account-mapping/sync', methods=['POST'])
def api_sync_account_mapping():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400
        mapping_list = data.get('mapping', [])
        if not mapping_list:
            return jsonify({'error': 'Пустой маппинг'}), 400
        result = db.sync_account_mapping_from_parser(mapping_list)
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API account-mapping/sync: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/delete-swipe', methods=['POST'])
def api_delete_swipe():
    try:
        data = request.get_json()
        swipe_date = data.get('date')
        if not swipe_date:
            return jsonify({'error': 'Не указана дата'}), 400
        result = db.delete_swipe_data(swipe_date)
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API delete-swipe: {e}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# API: БИБЛИОТЕКА КОНТРАГЕНТОВ (CRUD)
# ============================================================

@app.route('/api/contractors', methods=['GET'])
def api_get_contractors():
    try:
        contractors = db.get_all_contractors()
        return jsonify({'success': True, 'data': contractors})
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors GET: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors', methods=['POST'])
def api_add_contractor():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'error': 'Имя контрагента обязательно'}), 400
        organization = data.get('organization', '')
        explanation = data.get('explanation', '')
        result = db.insert_or_update_contractor(name, organization, explanation)
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors POST: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors/<int:contractor_id>', methods=['PUT'])
def api_update_contractor(contractor_id):
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400
        result = db.update_contractor(
            contractor_id,
            name=data.get('name'),
            organization=data.get('organization'),
            explanation=data.get('explanation')
        )
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors PUT: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors/<int:contractor_id>', methods=['DELETE'])
def api_delete_contractor(contractor_id):
    try:
        result = db.delete_contractor(contractor_id)
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors DELETE: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors/find', methods=['GET'])
def api_find_contractor():
    try:
        name = request.args.get('name', '').strip()
        if not name:
            return jsonify({'error': 'Не указано имя для поиска'}), 400
        result = db.find_contractor_by_name(name)
        if result:
            return jsonify({'success': True, 'data': result})
        else:
            return jsonify({'success': True, 'data': None})
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors/find: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors/import', methods=['POST'])
def api_import_contractors():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Нет данных'}), 400
        contractors_list = data.get('contractors', [])
        if not contractors_list:
            return jsonify({'error': 'Пустой список контрагентов'}), 400
        result = db.import_contractors_batch(contractors_list)
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors/import: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors/clear', methods=['POST'])
def api_clear_contractors():
    try:
        result = db.delete_all_contractors()
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors/clear: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/contractors/stats', methods=['GET'])
def api_contractors_stats():
    try:
        stats = db.get_contractors_stats()
        return jsonify({'success': True, 'data': stats})
    except Exception as e:
        logger.error(f"❌ Ошибка API contractors/stats: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================
# API: ГИБКИЕ НАСТРОЙКИ (правила исключений, категоризации, синонимы)
# ============================================================

@app.route('/api/config', methods=['GET'])
def api_get_config():
    try:
        config = db.get_all_config()
        return jsonify({'success': True, 'data': config})
    except Exception as e:
        logger.error(f"❌ Ошибка загрузки конфигурации: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exclusion-rules', methods=['GET'])
def api_get_exclusion_rules():
    try:
        rules = db.get_exclusion_rules()
        return jsonify({'success': True, 'data': rules})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exclusion-rules', methods=['POST'])
def api_add_exclusion_rule():
    try:
        data = request.get_json()
        result = db.save_exclusion_rule(
            data.get('rule_type', 'purpose'),
            data.get('pattern', ''),
            data.get('is_regex', False)
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exclusion-rules/<int:rule_id>', methods=['DELETE'])
def api_delete_exclusion_rule(rule_id):
    try:
        result = db.delete_exclusion_rule(rule_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exclusion-rules/<int:rule_id>', methods=['PUT'])
def api_update_exclusion_rule(rule_id):
    try:
        data = request.get_json()
        result = db.update_exclusion_rule(
            rule_id,
            rule_type=data.get('rule_type'),
            pattern=data.get('pattern'),
            is_regex=data.get('is_regex')
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categorization-rules', methods=['GET'])
def api_get_categorization_rules():
    try:
        rules = db.get_categorization_rules()
        return jsonify({'success': True, 'data': rules})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categorization-rules', methods=['POST'])
def api_add_categorization_rule():
    try:
        data = request.get_json()
        result = db.save_categorization_rule(
            data.get('field', 'purpose'),
            data.get('pattern', ''),
            data.get('display_name', '')
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categorization-rules/<int:rule_id>', methods=['DELETE'])
def api_delete_categorization_rule(rule_id):
    try:
        result = db.delete_categorization_rule(rule_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categorization-rules/<int:rule_id>', methods=['PUT'])
def api_update_categorization_rule(rule_id):
    try:
        data = request.get_json()
        result = db.update_categorization_rule(
            rule_id,
            field=data.get('field'),
            pattern=data.get('pattern'),
            display_name=data.get('display_name')
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/company-aliases', methods=['GET'])
def api_get_company_aliases():
    try:
        aliases = db.get_company_aliases()
        return jsonify({'success': True, 'data': aliases})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/company-aliases', methods=['POST'])
def api_add_company_alias():
    try:
        data = request.get_json()
        result = db.save_company_alias(
            data.get('pattern', ''),
            data.get('canonical_name', ''),
            data.get('match_type', 'contains')
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/company-aliases/<int:alias_id>', methods=['DELETE'])
def api_delete_company_alias(alias_id):
    try:
        result = db.delete_company_alias(alias_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/company-aliases/<int:alias_id>', methods=['PUT'])
def api_update_company_alias(alias_id):
    try:
        data = request.get_json()
        result = db.update_company_alias(
            alias_id,
            pattern=data.get('pattern'),
            canonical=data.get('canonical_name'),
            match_type=data.get('match_type')
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/account-mapping/<int:mapping_id>/toggle', methods=['PUT'])
def api_toggle_account(mapping_id):
    try:
        accounts = db.get_account_mapping(active_only=False)
        target = next((a for a in accounts if a['id'] == mapping_id), None)
        if not target:
            return jsonify({'success': False, 'error': 'Счёт не найден'}), 404
        new_active = not target['is_active']
        result = db.update_account_mapping(mapping_id, is_active=new_active)
        if result['success']:
            result['is_active'] = new_active
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================
# API: СРОЧНЫЕ ДЕПОЗИТЫ (term_deposits)
# ============================================================

@app.route('/api/term-deposits', methods=['GET'])
def api_get_term_deposits():
    """Получить все срочные депозиты."""
    try:
        deposits = db.get_term_deposits()
        # Группируем по account_number для фронтенда
        result = {}
        for d in deposits:
            acc = d['account_number']
            if acc not in result:
                result[acc] = []
            result[acc].append({
                'amount': float(d['amount']),
                'rate': float(d['rate']),
                'startDate': d['start_date'].isoformat() if hasattr(d['start_date'], 'isoformat') else str(d['start_date']),
                'endDate': d['end_date'].isoformat() if hasattr(d['end_date'], 'isoformat') else str(d['end_date'])
            })
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f"Ошибка API term-deposits GET: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/term-deposits/sync', methods=['POST'])
def api_sync_term_deposits():
    """Синхронизировать срочные депозиты (полная замена)."""
    try:
        data = request.get_json()
        deposits = data.get('deposits', {})
        result = db.sync_term_deposits(deposits)
        return jsonify({'success': True, 'synced': result})
    except Exception as e:
        logger.error(f"Ошибка API term-deposits/sync: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    logger.info("Сервер запущен на http://0.0.0.0:5000")
    try:
        app.run(debug=False, port=5000, host='0.0.0.0')
    finally:
        db.close_all_connections()
        logger.info("Сервер остановлен")