# allocator.py - Модуль распределения налогов по подразделениям с сохранением полной детализации
import openpyxl
from openpyxl.styles import numbers, Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import io
from collections import defaultdict

# ============================================
# КАТЕГОРИИ ДДС
# ============================================

REVENUE_CATEGORIES = {
    'Поступления от клиентов': [
        'Поступления от реализации товаров и услуг',
        'Поступления от клиентов',
    ],
    'Прочие поступления 4119': [
        'Прочие поступления 4119',
    ],
    'Возврат денежных средств от поставщика 4121': [
        'Возврат денежных средств от поставщика 4121',
    ],
    'Оплата поставщикам Прочие поставщики 4121': [
        'Оплата поставщикам Прочие поставщики 4121',
    ],
    'Поступления от аренды 4112': [
        'Поступления от аренды 4112',
    ],
    'Возврат, штрафы, возмещение страховок': [
        'Возврат денежных средств подотчетным лицом 4119',
        'Выплата заработной платы 4122',
        'Оплата поставщикам Страхование 4121',
        'Оплата за государственные пошлины, штрафы, исполнительные производства 4129',
    ],
    'Депозиты и %% по ним': [
        'Поступление процентов по депозитам 4119',
        'Возврат денежных средств с депозитов',
    ],
    'Продажа ТС': [
        'Выручка от продажи объектов основных средств и иных внеоборотных активов 4211',
    ],
    'Кредиты, займы и %%': [
        'Погашение займа, предоставленного юр.лицу 4213',
        'Получение займа от юр.лица 4311',
        'Поступление кредитов на пополнение оборотных средств 4311',
        'Поступления % по займу, предоставленного юр.лицу 4214',
        'Поступление инвестиционных кредитов 4311',
        'Сдача денежных средств в банк (в рублях)',
        'Оплата комиссии за неиспользованный лимит кредитной линии 4329',
    ],
    'Переброски': [
        'ВГО',
        'Перечисление денежных средств на другой счет',
    ],
}

EXPENSE_CATEGORIES = {
    'Прочие оплаты': [
        'Выплата заработной платы 4122',
        'Выдача денежных средств подотчетному лицу 4129',
        'Оплата за государственные пошлины, штрафы, исполнительные производства 4129',
        'Оплата за РКО 4129',
        'Прочие списания 4129',
    ],
    'Кредиты, займы и %%': [
        'Погашение кредитов на пополнение оборотных средств 4323',
        'Выплата процентов по кредитам на пополнение оборотных средств 4123',
        'Погашение инвестиционных кредитов 4323',
        'Выплата процентов по инвестиционным кредитам 4224',
        'Лизинговые платежи 4121',
        'Оплата комиссий за выдачу банковской гарантии 4129',
        'Оплата комиссии за неиспользованный лимит кредитной линии 4329',
        'Оплата комиссии за открытие и обслуживание ссудного счета 4129',
        'Оплата комиссий за выдачу кредита 4129',
        'Оплата комиссий за хеджирование',
        'Выдача займа юр.лицу 4223',
        'Погашение займа, полученного от юр.лица 4323',
        'Погашение % по займу, полученного от юр.лица 4123',
    ],
    'Налоги': [
        'Расчеты по налогам и сборам Налоги с ФОТ 4122',
        'Расчеты по налогам и сборам НДФЛ 4122',
        'Расчеты по налогам и сборам НДС 4129',
        'Расчеты по налогам и сборам Налог на прибыль 4124',
        'Расчеты по налогам и сборам прочие налоги и сборы 4129',
        'Расчеты по налогам и сборам Пени штрафы 4129',
    ],
    'Поставщики': [
        'Оплата поставщикам Прочие поставщики 4121',
        'Оплата поставщикам Субподрядчики 4121',
        'Оплата поставщикам Транспортные услуги 4121',
        'Оплата поставщикам Аренда ТС 4121',
        'Оплата поставщикам Запчасти, ремонт 4121',
        'Оплата поставщикам Шины 4121',
        'Оплата поставщикам Спецодежда 4121',
        'Оплата поставщикам Аренда помещений, коммунальные платежи и связь 4121',
        'Оплата поставщикам Питание 4121',
        'Оплата за топливо 4121',
        'Оплата масел и иных смазочных материалов 4121',
        'Оплата поставщикам Перелеты и тревел-поставщики 4121',
        'Возврат оплаты покупателю, не аренда 4111',
        'Возврат оплаты покупателю, аренда 4112',
    ],
    'Страхование': [
        'Оплата поставщикам Страхование 4121',
    ],
    'Депозиты и %%': [
        'Размещение денежных средств на депозиты',
    ],
    'Инвестиции': [
        'Инвестиции в ТС 4221',
        'Инвестиции в инфраструктуру и формирование ОС 4221',
    ],
    'Переброски': [
        'ВГО',
        'Перечисление денежных средств на другой счет',
    ],
}

# Статьи, которые относятся к налогам и подлежат распределению
TAX_ARTICLES = [
    'Расчеты по налогам и сборам Налоги с ФОТ 4122',
    'Расчеты по налогам и сборам НДФЛ 4122',
    'Расчеты по налогам и сборам НДС 4129',
    'Расчеты по налогам и сборам Налог на прибыль 4124',
]

# ============================================
# ФОРМАТИРОВАНИЕ (ЦВЕТА И ШРИФТЫ)
# ============================================

# Цвета фона
HEADER_BLUE_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
CATEGORY_HEADER_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_ROW_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Шрифты
FONT_COMPANY = Font(bold=True, size=14, color='000000')
FONT_SECTION = Font(bold=True, size=12, color='000000')
FONT_CATEGORY = Font(bold=True, size=11, color='000000')
FONT_REGULAR = Font(bold=False, size=11, color='000000')
FONT_TOTAL = Font(bold=True, size=11, color='000000')
FONT_HEADER = Font(bold=True, size=11, color='FFFFFF')

# Границы
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)


class CashFlowAllocator:
    """Класс для распределения налогов по подразделениям с сохранением полной детализации"""
    
    def __init__(self, strategy='revenue', rounding='kopeks'):
        self.strategy = strategy
        self.rounding = rounding
        self.original_rows = []
        self.filials = {}
        self.taxes = []
        self.shares = {}
        self.total_revenue = 0
        self.months = set()
        self.headers = []
        self.company_name = 'СОИР ООО'
        self.opening_balance = 0
        
    def load_excel(self, file_bytes):
        """Загружает Excel файл с сохранением ВСЕХ колонок
        
        data_only=True — читаем вычисленные значения формул (например,
        столбец 'Месяц оплаты' содержит =TEXT(G10,"ММММ.ГГ") вместо текста).
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        if 'Источник' not in wb.sheetnames:
            raise ValueError("Файл не содержит вкладку 'Источник'")
        
        ws = wb['Источник']
        self.original_rows = []
        
        # Сохраняем заголовки
        self.headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            self.headers.append(str(cell_value).strip() if cell_value else '')
        
        print(f"📊 Заголовки в файле: {self.headers[:10]}")
        
        # Определяем индексы колонок
        col_indices = {}
        for idx, header in enumerate(self.headers):
            header_lower = header.lower() if header else ''
            
            if 'подразделение' in header_lower or 'филиал' in header_lower:
                col_indices['подразделение'] = idx
            elif 'статья движения' in header_lower or 'статья' in header_lower or 'назначение' in header_lower:
                col_indices['статья'] = idx
            elif 'факт' in header_lower or 'сумма' in header_lower or 'прогноз' in header_lower:
                col_indices['сумма'] = idx
            elif 'месяц' in header_lower or 'период' in header_lower:
                col_indices['месяц'] = idx
            elif 'раздел' in header_lower:
                col_indices['раздел'] = idx
            elif 'организация' in header_lower:
                col_indices['организация'] = idx
                # Извлекаем название компании
                if idx < ws.max_column:
                    for row in range(2, min(ws.max_row + 1, 10)):
                        org_value = ws.cell(row=row, column=idx + 1).value
                        if org_value and str(org_value).strip():
                            self.company_name = str(org_value).strip()
                            break
            elif 'контрагент' in header_lower:
                col_indices['контрагент'] = idx
            elif 'договор' in header_lower:
                col_indices['договор'] = idx
            elif 'дата' in header_lower:
                col_indices['дата'] = idx
        
        print(f"📋 Найдены индексы колонок: {col_indices}")
        print(f"🏢 Название компании: {self.company_name}")
        
        # Читаем ВСЕ строки
        for row in range(2, ws.max_row + 1):
            try:
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                
                подразделение_idx = col_indices.get('подразделение', 1)
                статья_idx = col_indices.get('статья', 5)
                сумма_idx = col_indices.get('сумма', 7)
                месяц_idx = col_indices.get('месяц', 8)
                
                подразделение = row_data[подразделение_idx] if подразделение_idx < len(row_data) else None
                статья = row_data[статья_idx] if статья_idx < len(row_data) else None
                сумма = row_data[сумма_idx] if сумма_idx < len(row_data) else None
                месяц = row_data[месяц_idx] if месяц_idx < len(row_data) else None
                
                if not статья:
                    continue
                
                месяц_str = str(месяц).strip() if месяц else ''
                
                # ✅ Если месяц — формула (начинается с '=') или пустой — вычисляем из даты
                if not месяц_str or месяц_str.startswith('='):
                    дата_idx = col_indices.get('дата', 6)
                    дата_знач = row_data[дата_idx] if дата_idx < len(row_data) else None
                    месяц_str = self._date_to_month(дата_знач)
                
                self.original_rows.append({
                    'raw_data': row_data,
                    'подразделение': str(подразделение).strip() if подразделение else '',
                    'статья': str(статья).strip() if статья else '',
                    'сумма': float(сумма) if сумма is not None and сумма != "" else 0,
                    'месяц': месяц_str
                })
                
                if месяц_str:
                    self.months.add(месяц_str)
                    
            except Exception as e:
                print(f"⚠️ Ошибка строки {row}: {e}")
                continue
        
        print(f"✅ Загружено {len(self.original_rows)} записей, месяцев: {len(self.months)}")
        return len(self.original_rows)
    
    def is_tax_article(self, article):
        """Проверяет, является ли статья налоговой"""
        if not article:
            return False
        article_lower = article.lower()
        for tax_article in TAX_ARTICLES:
            if tax_article.lower() in article_lower or article_lower in tax_article.lower():
                return True
        return False
    
    def is_revenue_article(self, article):
        """Проверяет, является ли статья выручкой"""
        if not article:
            return False
        article_lower = article.lower()
        return any(kw in article_lower for kw in ['поступления от реализации', 'поступления от клиентов'])
    
    def find_category(self, article):
        """Находит категорию для статьи"""
        if not article:
            return None, None
        
        article_lower = article.lower()
        
        # Сначала ищем в расходах
        for category_name, templates in EXPENSE_CATEGORIES.items():
            for template in templates:
                template_lower = template.lower()
                if template_lower in article_lower or article_lower in template_lower:
                    return category_name, template
        
        # Потом в доходах
        for category_name, templates in REVENUE_CATEGORIES.items():
            for template in templates:
                template_lower = template.lower()
                if template_lower in article_lower or article_lower in template_lower:
                    return category_name, template
        
        return None, None
    
    def separate_data(self):
        """Разделяет данные по подразделениям и месяцам"""
        self.filials = defaultdict(lambda: defaultdict(list))
        self.taxes = []
        self.total_revenue = 0
        
        for row in self.original_rows:
            подразделение = row['подразделение']
            статья = row['статья']
            сумма = row['сумма']
            месяц = row['месяц']
            
            category, template = self.find_category(статья)
            
            if подразделение:
                self.filials[подразделение][месяц].append({
                    **row,
                    'category': category,
                    'template': template
                })
            
            if сумма > 0 and self.is_revenue_article(статья):
                self.total_revenue += сумма
            
            if self.is_tax_article(статья):
                self.taxes.append(row)
        
        print(f"✅ Найдено налогов: {len(self.taxes)}, Подразделений: {len(self.filials)}")
        return len(self.taxes)
    
    def calculate_shares(self):
        """Рассчитывает доли подразделений в общей выручке"""
        filial_revenue = defaultdict(float)
        
        for filial, months_data in self.filials.items():
            for month, rows in months_data.items():
                for row in rows:
                    if row['сумма'] > 0 and self.is_revenue_article(row['статья']):
                        filial_revenue[filial] += row['сумма']
        
        if self.total_revenue == 0:
            equal_share = 1.0 / len(self.filials) if self.filials else 0
            for filial in self.filials:
                self.shares[filial] = equal_share
            return self.shares
        
        for filial, revenue in filial_revenue.items():
            self.shares[filial] = revenue / self.total_revenue
        
        print(f"💰 Доли филиалов рассчитаны: {len(self.shares)} филиалов")
        return self.shares
    
    def _get_tax_type(self, article):
        """Определяет тип налога по статье"""
        article_lower = article.lower()
        if 'ндфл' in article_lower:
            return 'НДФЛ'
        elif 'налоги с фот' in article_lower:
            return 'Налоги с ФОТ'
        elif 'ндс' in article_lower:
            return 'НДС'
        elif 'прибыль' in article_lower:
            return 'Налог на прибыль'
        elif 'пени' in article_lower or 'штраф' in article_lower:
            return 'Пени штрафы'
        else:
            return 'Прочие налоги'
    
    def allocate_taxes(self):
        """Распределяет налоги по подразделениям"""
        allocated_taxes = []
        
        tax_groups = defaultdict(lambda: {'total': 0, 'месяц': '', 'статья': '', 'raw_template': None})
        
        for tax in self.taxes:
            tax_type = self._get_tax_type(tax['статья'])
            месяц = tax['месяц'] or 'Без месяца'
            key = (tax_type, месяц)
            tax_groups[key]['total'] += abs(tax['сумма'])
            tax_groups[key]['месяц'] = месяц
            tax_groups[key]['статья'] = tax['статья']
            if tax_groups[key]['raw_template'] is None:
                tax_groups[key]['raw_template'] = tax
        
        for (tax_type, месяц), tax_data in tax_groups.items():
            total_tax = tax_data['total']
            original_tax_row = tax_data['raw_template']
            
            for filial, share in self.shares.items():
                allocated_amount = total_tax * share
                if self.rounding == 'rubles':
                    allocated_amount = round(allocated_amount)
                else:
                    allocated_amount = round(allocated_amount, 2)
                
                new_row_data = None
                if original_tax_row and 'raw_data' in original_tax_row:
                    new_row_data = original_tax_row['raw_data'].copy()
                    подразделение_idx = self._get_column_index('подразделение')
                    сумма_idx = self._get_column_index('Факт') or self._get_column_index('сумма')
                    if подразделение_idx is not None and подразделение_idx < len(new_row_data):
                        new_row_data[подразделение_idx] = filial
                    if сумма_idx is not None and сумма_idx < len(new_row_data):
                        new_row_data[сумма_idx] = -allocated_amount
                else:
                    new_row_data = [''] * len(self.headers)
                    for idx, header in enumerate(self.headers):
                        header_lower = header.lower() if header else ''
                        if 'подразделение' in header_lower:
                            new_row_data[idx] = filial
                        elif 'статья' in header_lower:
                            new_row_data[idx] = tax_data['статья']
                        elif 'сумма' in header_lower or 'факт' in header_lower:
                            new_row_data[idx] = -allocated_amount
                        elif 'месяц' in header_lower:
                            new_row_data[idx] = месяц
                
                allocated_taxes.append({
                    'raw_data': new_row_data,
                    'подразделение': filial,
                    'статья': tax_data['статья'],
                    'сумма': -allocated_amount,
                    'месяц': месяц,
                    'тип_налога': tax_type,
                    'category': 'Налоги'
                })
        
        print(f"✅ Распределено {len(allocated_taxes)} налоговых записей")
        return allocated_taxes
    
    def _date_to_month(self, date_value):
        """Вычисляет месяц в формате 'Апрель.26' из значения даты."""
        if not date_value:
            return ''
        month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                       'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        if isinstance(date_value, datetime):
            return f"{month_names[date_value.month - 1]}.{str(date_value.year)[-2:]}"
        if isinstance(date_value, str):
            for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%m/%d/%Y'):
                try:
                    dt = datetime.strptime(date_value.strip(), fmt)
                    return f"{month_names[dt.month - 1]}.{str(dt.year)[-2:]}"
                except ValueError:
                    continue
        return ''

    def _get_column_index(self, column_name):
        """Возвращает индекс колонки по имени"""
        for idx, header in enumerate(self.headers):
            header_lower = header.lower() if header else ''
            if column_name in header_lower:
                return idx
        return None
    
    def generate_excel(self, output_path=None):
        """Генерирует Excel файл"""
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        allocated_taxes = self.allocate_taxes()
        
        all_rows_for_source = []
        
        for row in self.original_rows:
            if not self.is_tax_article(row['статья']):
                all_rows_for_source.append(row['raw_data'])
        
        for tax_row in allocated_taxes:
            if tax_row['raw_data']:
                all_rows_for_source.append(tax_row['raw_data'])
        
        source_ws = wb.create_sheet('Источник')
        self._write_source_sheet_with_full_data(source_ws, all_rows_for_source)
        
        self._prepare_filial_data_with_taxes(allocated_taxes)
        
        for filial in sorted(self.filials.keys()):
            ws = wb.create_sheet(filial)
            self._write_filial_dds_sheet(ws, filial)
        
        if output_path:
            wb.save(output_path)
            print(f"✅ Файл сохранён: {output_path}")
        
        return wb
    
    def _write_source_sheet_with_full_data(self, ws, all_rows):
        """Записывает лист Источник"""
        for col, header in enumerate(self.headers, 1):
            if header:
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = FONT_HEADER
                cell.fill = HEADER_BLUE_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = THIN_BORDER
        
        for row_idx, row_data in enumerate(all_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                if value is not None:
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    fact_idx = self._get_column_index('Факт')
                    summa_idx = self._get_column_index('сумма')
                    if (fact_idx is not None and col_idx == fact_idx + 1) or (summa_idx is not None and col_idx == summa_idx + 1):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = THIN_BORDER
        
        print(f"✅ Лист 'Источник' создан: {len(all_rows)} строк")
    
    def _prepare_filial_data_with_taxes(self, allocated_taxes):
        """Добавляет распределённые налоги в данные филиалов"""
        for filial in self.filials:
            for month in self.filials[filial]:
                self.filials[filial][month] = [
                    row for row in self.filials[filial][month]
                    if not self.is_tax_article(row.get('статья', ''))
                ]
        
        for tax in allocated_taxes:
            filial = tax['подразделение']
            month = tax['месяц']
            if filial not in self.filials:
                self.filials[filial] = defaultdict(list)
            self.filials[filial][month].append(tax)
    
    def _write_filial_dds_sheet(self, ws, filial_name):
        """Записывает ДДС по подразделению — ОДНА КОЛОНКА за весь период"""
        months_data = self.filials[filial_name]
        
        print(f"📝 Создаём ДДС для {filial_name}")
        
        # Агрегируем данные — суммируем за весь период, без разбивки по месяцам
        income_data = defaultdict(float)   # статья -> общая сумма (положительные)
        expense_data = defaultdict(float)  # статья -> общая сумма (отрицательные)
        
        for month, rows in months_data.items():
            for row in rows:
                article = row['статья']
                amount = row['сумма']
                if amount > 0:
                    income_data[article] += amount
                elif amount < 0:
                    expense_data[article] += amount
        
        TOTAL_COL = 3  # A=статья, B=Сумма, C=Итого
        
        # Строка 1: Название подразделения
        ws.cell(row=1, column=1, value=filial_name)
        ws.cell(row=1, column=1).font = FONT_COMPANY
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COL)
        
        # Строка 2: Компания
        ws.cell(row=2, column=1, value=self.company_name)
        ws.cell(row=2, column=1).font = FONT_REGULAR
        
        # Строка 3: Входящий остаток
        ws.cell(row=3, column=1, value='Входящий остаток')
        ws.cell(row=3, column=1).font = FONT_REGULAR
        ws.cell(row=3, column=2).font = FONT_REGULAR
        ws.cell(row=3, column=2).number_format = '#,##0.00'
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='right')
        
        # Строка 4: Заголовки
        ws.cell(row=4, column=1, value='Период')
        ws.cell(row=4, column=1).font = FONT_REGULAR
        ws.cell(row=4, column=2, value='Сумма').font = FONT_HEADER
        ws.cell(row=4, column=2).fill = HEADER_BLUE_FILL
        ws.cell(row=4, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=4, column=2).border = THIN_BORDER
        ws.cell(row=4, column=3, value='Итого').font = FONT_HEADER
        ws.cell(row=4, column=3).fill = HEADER_BLUE_FILL
        ws.cell(row=4, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=4, column=3).border = THIN_BORDER
        
        row = 5
        
        # ===== ПОСТУПЛЕНИЯ =====
        ws.cell(row=row, column=1, value='Поступления')
        ws.cell(row=row, column=1).font = FONT_SECTION
        for col in range(1, TOTAL_COL + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        
        total_income = 0
        for category_name, articles in REVENUE_CATEGORIES.items():
            row, cat_total = self._write_category_row_simple(ws, row, income_data, category_name, articles, TOTAL_COL)
            total_income += cat_total
        
        # Итого поступления
        ws.cell(row=row, column=1, value='Итого поступления')
        ws.cell(row=row, column=1).font = FONT_TOTAL
        ws.cell(row=row, column=1).fill = TOTAL_ROW_FILL
        self._write_total_cell(ws, row, 2, total_income, TOTAL_ROW_FILL)
        self._write_total_cell(ws, row, 3, total_income, TOTAL_ROW_FILL)
        for col in range(1, TOTAL_COL + 1):
            ws.cell(row=row, column=col).fill = TOTAL_ROW_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 2
        
        # ===== РАСХОДЫ =====
        ws.cell(row=row, column=1, value='Расходы')
        ws.cell(row=row, column=1).font = FONT_SECTION
        for col in range(1, TOTAL_COL + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        
        total_expense = 0
        for category_name, articles in EXPENSE_CATEGORIES.items():
            row, cat_total = self._write_category_row_simple(ws, row, expense_data, category_name, articles, TOTAL_COL)
            total_expense += cat_total
        
        # Итого расходы
        ws.cell(row=row, column=1, value='Итого расходы')
        ws.cell(row=row, column=1).font = FONT_TOTAL
        ws.cell(row=row, column=1).fill = TOTAL_ROW_FILL
        self._write_total_cell(ws, row, 2, total_expense, TOTAL_ROW_FILL)
        self._write_total_cell(ws, row, 3, total_expense, TOTAL_ROW_FILL)
        for col in range(1, TOTAL_COL + 1):
            ws.cell(row=row, column=col).fill = TOTAL_ROW_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 2
        
        # Исходящий остаток
        closing = self.opening_balance + total_income + total_expense
        ws.cell(row=row, column=1, value='Исходящий остаток')
        ws.cell(row=row, column=1).font = FONT_TOTAL
        ws.cell(row=row, column=1).fill = TOTAL_ROW_FILL
        self._write_total_cell(ws, row, 2, closing, TOTAL_ROW_FILL)
        self._write_total_cell(ws, row, 3, closing, TOTAL_ROW_FILL)
        for col in range(1, TOTAL_COL + 1):
            ws.cell(row=row, column=col).fill = TOTAL_ROW_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        
        ws.sheet_view.showOutlineSymbols = True
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
    
    def _write_category_row_simple(self, ws, row, data, category_name, articles, TOTAL_COL):
        """Записывает категорию доходов/расходов — одна колонка Сумма"""
        category_total = 0
        has_data = False
        
        for article in articles:
            amount = data.get(article, 0)
            if amount != 0:
                category_total += amount
                has_data = True
        
        # Заголовок категории (жёлтый)
        ws.cell(row=row, column=1, value=category_name)
        ws.cell(row=row, column=1).font = FONT_CATEGORY
        ws.cell(row=row, column=1).fill = CATEGORY_HEADER_FILL
        
        if has_data:
            self._write_total_cell(ws, row, 2, category_total, CATEGORY_HEADER_FILL)
            self._write_total_cell(ws, row, 3, category_total, CATEGORY_HEADER_FILL)
        else:
            for col in range(1, TOTAL_COL + 1):
                ws.cell(row=row, column=col).fill = CATEGORY_HEADER_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER
        
        for col in range(1, TOTAL_COL + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        
        row += 1
        
        # Статьи внутри категории
        for article in articles:
            amount = data.get(article, 0)
            if amount != 0:
                ws.cell(row=row, column=1, value=article)
                ws.cell(row=row, column=1).font = FONT_REGULAR
                self._write_total_cell(ws, row, 2, amount)
                self._write_total_cell(ws, row, 3, amount)
                for col in range(1, TOTAL_COL + 1):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1
        
        return row, category_total
    
    def _write_total_cell(self, ws, row, col, value, fill=None):
        """Записывает числовую ячейку с форматированием"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.font = FONT_TOTAL
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    
    def _prepare_dds_data(self, months_data, sorted_months):
        """Подготавливает данные для ДДС"""
        dds_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
        
        for month in sorted_months:
            rows = months_data.get(month, [])
            for row in rows:
                article = row['статья']
                amount = row['сумма']
                category, template = self.find_category(article)
                if category:
                    dds_data[month][category][article] += amount
        
        return dds_data


def allocate_cashflow(file_bytes, strategy='revenue', rounding='kopeks', 
                      period_type='month', period_value='', 
                      date_from='', date_to='', exclude_articles=None):
    """Основная функция для распределения налогов и формирования ДДС"""
    allocator = CashFlowAllocator(strategy=strategy, rounding=rounding)
    
    allocator.load_excel(file_bytes)
    
    # ✅ ФИЛЬТРАЦИЯ: удаляем строки не соответствующие выбранному периоду
    if period_type == 'month' and period_value:
        before = len(allocator.original_rows)
        allocator.original_rows = [
            row for row in allocator.original_rows
            if row.get('месяц', '') == period_value
        ]
        after = len(allocator.original_rows)
        print(f"🔍 Отфильтровано по месяцу '{period_value}': {after} из {before} строк")
        # Если после фильтрации не осталось данных — возвращаем ошибку
        if after == 0:
            return {
                'success': False,
                'error': f"После фильтрации по месяцу '{period_value}' не осталось данных. "
                         f"Доступные месяцы в файле: {sorted(allocator.months)}"
            }
    
    # ✅ ИСКЛЮЧЕНИЕ: удаляем строки с исключёнными статьями
    if exclude_articles:
        before = len(allocator.original_rows)
        allocator.original_rows = [
            row for row in allocator.original_rows
            if not any(ex.lower() in str(row.get('статья', '')).lower() for ex in exclude_articles)
        ]
        after = len(allocator.original_rows)
        print(f"🔍 Исключено статей: {before - after} строк (из {before})")
    
    allocator.separate_data()
    allocator.calculate_shares()
    
    output = io.BytesIO()
    allocator.generate_excel(output)
    output.seek(0)
    
    return {
        'success': True,
        'file': output,
        'filials_count': len(allocator.filials),
        'total_revenue': allocator.total_revenue,
        'total_taxes': sum(abs(t['сумма']) for t in allocator.taxes),
        'shares': allocator.shares
    }


if __name__ == '__main__':
    print("Модуль распределения налогов загружен")
